"""考核打分系统 API"""
from flask import Blueprint, request, jsonify, send_file
from supabase_config import get_db
from psycopg2.extras import RealDictCursor
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

bp = Blueprint('evaluation', __name__)

# ============= 角色判定 =============

def get_evaluator_role(person):
    """根据人员岗位判断打分角色"""
    if not person:
        return None
    position = (person.get('position') or '').lower()
    dept = person.get('dept') or ''

    # 领导班子：经理、书记、商务经理
    if any(k in position for k in ['经理', '书记']):
        return 'leader'

    # 部门负责人：部长、副部长、主任、负责人、主管
    if any(k in position for k in ['部长', '主任', '负责人', '主管', '副主管']):
        return 'dept_leader'

    # 普通员工（互评）
    return 'peer'


def get_my_evaluatees(cur, evaluator_phone):
    """根据打分人角色动态获取被考核人列表

    - 领导班子 → 所有一般员工（peer，即不含领导班子/部门负责人）
    - 部门负责人 → 本部门所有一般员工
    - 普通员工 → 其他所有普通员工
    """
    cur.execute("""
        SELECT id, name, dept, project, position, category, phone, leave_date
        FROM personnel
        WHERE (leave_date IS NULL OR leave_date = '')
        ORDER BY category, position, name
    """)
    all_personnel = cur.fetchall()

    # 找到当前用户
    me = None
    for p in all_personnel:
        if p.get('phone') == evaluator_phone:
            me = p
            break
    if not me:
        return [], None

    my_role = get_evaluator_role(me)
    if not my_role:
        return [], None

    my_id = me['id']
    my_dept = me.get('dept') or ''
    my_position = me.get('position') or ''

    # 过滤掉领导班子和部门负责人（即只保留一般员工作为被考核人）
    def is_leader_or_dept_leader(p):
        pos = (p.get('position') or '').lower()
        return any(k in pos for k in ['经理', '书记', '部长', '主任', '负责人', '主管', '副主管'])

    evaluatees = []

    if my_role == 'leader':
        # 领导班子：所有一般员工
        evaluatees = [p for p in all_personnel
                      if p['id'] != my_id and not is_leader_or_dept_leader(p)]

    elif my_role == 'dept_leader':
        # 部门负责人：本部门一般员工
        evaluatees = [p for p in all_personnel
                      if p['id'] != my_id
                      and p.get('dept') == my_dept
                      and not is_leader_or_dept_leader(p)]
        # 如果本部门没人，回退到所有一般员工
        if not evaluatees:
            evaluatees = [p for p in all_personnel
                          if p['id'] != my_id and not is_leader_or_dept_leader(p)]

    else:
        # 普通员工：其他所有普通员工（互评）
        evaluatees = [p for p in all_personnel
                      if p['id'] != my_id and not is_leader_or_dept_leader(p)]

    return evaluatees, my_role


def get_evaluator_name(cur, phone):
    """根据手机号获取打分人姓名"""
    cur.execute("SELECT name FROM users WHERE phone = %s", (phone,))
    r = cur.fetchone()
    if r:
        return r['name']
    # fallback: 从personnel表查
    cur.execute("SELECT name FROM personnel WHERE phone = %s LIMIT 1", (phone,))
    r = cur.fetchone()
    return r['name'] if r else phone


def is_admin(user):
    """判断是否是超级管理员"""
    if not user:
        return False
    return user.get('is_admin') == True or user.get('phone') == '18184005669'


# ============= API =============

@bp.route('/api/evaluation/cycles', methods=['GET'])
def list_cycles():
    """获取考核周期列表"""
    user = request.headers.get('X-User-Phone')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, period, cycle_type, status, start_date, end_date, created_at, created_by
        FROM evaluation_cycles
        ORDER BY id DESC
    """)
    cycles = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify({'cycles': cycles, 'user_phone': user})


@bp.route('/api/evaluation/cycles', methods=['POST'])
def create_cycle():
    """创建考核周期（仅超管）"""
    user_phone = request.headers.get('X-User-Phone', '')
    if user_phone != '18184005669':
        return jsonify({'error': '无权限'}), 403
    data = request.json
    name = data.get('name', '').strip()
    period = data.get('period', '').strip()
    if not name or not period:
        return jsonify({'error': '周期名称和编号必填'}), 400

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO evaluation_cycles
            (name, period, cycle_type, status, start_date, end_date, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            name,
            period,
            data.get('cycle_type', 'quarter'),
            data.get('status', 'active'),
            data.get('start_date', ''),
            data.get('end_date', ''),
            user_phone
        ))
        new_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'success': True, 'id': new_id})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@bp.route('/api/evaluation/cycles/<int:cycle_id>/close', methods=['POST'])
def close_cycle(cycle_id):
    """关闭考核周期（仅超管）"""
    user_phone = request.headers.get('X-User-Phone', '')
    if user_phone != '18184005669':
        return jsonify({'error': '无权限'}), 403
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE evaluation_cycles SET status='closed' WHERE id=%s", (cycle_id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'success': True})


@bp.route('/api/evaluation/my-tasks', methods=['GET'])
def my_tasks():
    """获取我的打分任务（按角色自动生成）

    返回：
    - cycle: 当前周期
    - my_role: 我的打分角色
    - evaluatees: 待打分列表（带是否已打分标记）
    """
    user_phone = request.headers.get('X-User-Phone')
    if not user_phone:
        return jsonify({'error': '未登录'}), 401

    cycle_id = request.args.get('cycle_id', type=int)
    conn = get_db()
    cur = conn.cursor()

    # 取周期
    if cycle_id:
        cur.execute("SELECT * FROM evaluation_cycles WHERE id=%s", (cycle_id,))
    else:
        cur.execute("SELECT * FROM evaluation_cycles WHERE status='active' ORDER BY id DESC LIMIT 1")
    cycle = cur.fetchone()
    if not cycle:
        cur.close()
        conn.close()
        return jsonify({'error': '当前没有进行中的考核周期'}), 404

    # 取我的评分对象
    evaluatees, my_role = get_my_evaluatees(cur, user_phone)
    if not evaluatees or not my_role:
        cur.close()
        conn.close()
        return jsonify({
            'cycle': cycle,
            'my_role': my_role,
            'evaluatees': [],
            'evaluator_name': get_evaluator_name(cur, user_phone),
            'message': '当前角色没有需要打分的人员'
        })

    # 标记已打分
    cur.execute("""
        SELECT evaluatee_id, performance_score, ability_score, attitude_score,
               execution_score, discipline_score, total_score, comment, submitted_at
        FROM evaluation_scores
        WHERE cycle_id=%s AND evaluator_phone=%s
    """, (cycle['id'], user_phone))
    scored_map = {r['evaluatee_id']: r for r in cur.fetchall()}

    # 标记 + 拼接到列表
    result_list = []
    for p in evaluatees:
        sc = scored_map.get(p['id'])
        item = {
            'id': p['id'],
            'name': p['name'],
            'dept': p.get('dept', ''),
            'project': p.get('project', ''),
            'position': p.get('position', ''),
            'category': p.get('category', ''),
            'scored': sc is not None,
            'total_score': float(sc['total_score']) if sc else None,
        }
        if sc:
            item['submitted_at'] = sc['submitted_at'].strftime('%Y-%m-%d %H:%M') if sc['submitted_at'] else ''
        result_list.append(item)

    evaluator_name = get_evaluator_name(cur, user_phone)
    cur.close()
    conn.close()

    # 预计算百分比（WXML 不支持方法调用）
    total = len(result_list)
    scored = sum(1 for x in result_list if x['scored'])
    percent = round(scored / total * 100) if total > 0 else 0

    return jsonify({
        'cycle': cycle,
        'my_role': my_role,
        'evaluator_name': evaluator_name,
        'evaluatees': result_list,
        'total_count': total,
        'scored_count': scored,
        'percent': percent
    })


@bp.route('/api/evaluation/scores/<int:cycle_id>/<evaluatee_id>', methods=['GET'])
def get_score(cycle_id, evaluatee_id):
    """获取我对某人的具体打分（用于编辑）"""
    user_phone = request.headers.get('X-User-Phone')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT * FROM evaluation_scores
        WHERE cycle_id=%s AND evaluator_phone=%s AND evaluatee_id=%s
    """, (cycle_id, user_phone, evaluatee_id))
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return jsonify({'scored': False})
    return jsonify({
        'scored': True,
        'performance_score': float(row['performance_score']),
        'ability_score': float(row['ability_score']),
        'attitude_score': float(row['attitude_score']),
        'execution_score': float(row['execution_score']),
        'discipline_score': float(row['discipline_score']),
        'total_score': float(row['total_score']),
        'comment': row['comment'] or '',
    })


@bp.route('/api/evaluation/submit', methods=['POST'])
def submit_score():
    """提交打分（支持单条/批量）"""
    user_phone = request.headers.get('X-User-Phone')
    if not user_phone:
        return jsonify({'error': '未登录'}), 401

    data = request.json
    cycle_id = data.get('cycle_id')
    scores = data.get('scores')  # 兼容批量：[{evaluatee_id, performance_score, ...}, ...]
    if not scores and data.get('evaluatee_id'):
        scores = [data]

    if not cycle_id or not scores:
        return jsonify({'error': '参数缺失'}), 400

    conn = get_db()
    cur = conn.cursor()
    try:
        evaluator_name = get_evaluator_name(cur, user_phone)

        # 获取当前用户角色（用于决定权重，但存不存都可以）
        evaluatees, my_role = get_my_evaluatees(cur, user_phone)
        if not my_role:
            return jsonify({'error': '无打分权限'}), 403

        # 取被考核人信息缓存
        eval_id_to_info = {p['id']: p for p in evaluatees}

        success_count = 0
        for s in scores:
            eid = s.get('evaluatee_id')
            perf = float(s.get('performance_score', 0))
            abil = float(s.get('ability_score', 0))
            atti = float(s.get('attitude_score', 0))
            exec_ = float(s.get('execution_score', 0))
            disc = float(s.get('discipline_score', 0))
            total = perf + abil + atti + exec_ + disc
            comment = (s.get('comment') or '').strip()

            info = eval_id_to_info.get(eid, {})

            # UPSERT
            cur.execute("""
                INSERT INTO evaluation_scores
                (cycle_id, evaluator_phone, evaluator_name, evaluator_role,
                 evaluatee_id, evaluatee_name, evaluatee_dept, evaluatee_project,
                 performance_score, ability_score, attitude_score, execution_score, discipline_score,
                 total_score, comment, submitted_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                ON CONFLICT (cycle_id, evaluator_phone, evaluatee_id)
                DO UPDATE SET
                    performance_score = EXCLUDED.performance_score,
                    ability_score = EXCLUDED.ability_score,
                    attitude_score = EXCLUDED.attitude_score,
                    execution_score = EXCLUDED.execution_score,
                    discipline_score = EXCLUDED.discipline_score,
                    total_score = EXCLUDED.total_score,
                    comment = EXCLUDED.comment,
                    submitted_at = NOW()
            """, (
                cycle_id, user_phone, evaluator_name, my_role,
                eid, info.get('name', ''), info.get('dept', ''), info.get('project', ''),
                perf, abil, atti, exec_, disc,
                total, comment
            ))
            success_count += 1

        conn.commit()
        return jsonify({'success': True, 'count': success_count})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@bp.route('/api/evaluation/results', methods=['GET'])
def get_results():
    """查看考核汇总结果（仅超管）"""
    user_phone = request.headers.get('X-User-Phone', '')
    if user_phone != '18184005669':
        return jsonify({'error': '无权限'}), 403

    cycle_id = request.args.get('cycle_id', type=int)
    conn = get_db()
    cur = conn.cursor()

    if cycle_id:
        cur.execute("SELECT * FROM evaluation_cycles WHERE id=%s", (cycle_id,))
    else:
        cur.execute("SELECT * FROM evaluation_cycles ORDER BY id DESC LIMIT 1")
    cycle = cur.fetchone()
    if not cycle:
        cur.close()
        conn.close()
        return jsonify({'error': '还没有考核周期'}), 404

    # 按被考核人汇总：领导班子均分 + 部门负责人均分 + 员工互评均分 = 综合分
    cur.execute("""
        SELECT
            evaluatee_id,
            evaluatee_name,
            evaluatee_dept,
            evaluatee_project,
            evaluator_role,
            COUNT(*) as evaluator_count,
            ROUND(AVG(total_score)::numeric, 2) as avg_score
        FROM evaluation_scores
        WHERE cycle_id = %s
        GROUP BY evaluatee_id, evaluatee_name, evaluatee_dept, evaluatee_project, evaluator_role
    """, (cycle['id'],))
    grouped = cur.fetchall()

    # 重组：每个人一行
    by_person = {}
    for r in grouped:
        key = r['evaluatee_id']
        if key not in by_person:
            by_person[key] = {
                'id': r['evaluatee_id'],
                'name': r['evaluatee_name'],
                'dept': r['evaluatee_dept'] or '',
                'project': r['evaluatee_project'] or '',
                'leader_score': None, 'leader_count': 0,
                'dept_score': None, 'dept_count': 0,
                'peer_score': None, 'peer_count': 0,
            }
        role = r['evaluator_role']
        if role == 'leader':
            by_person[key]['leader_score'] = float(r['avg_score'])
            by_person[key]['leader_count'] = r['evaluator_count']
        elif role == 'dept_leader':
            by_person[key]['dept_score'] = float(r['avg_score'])
            by_person[key]['dept_count'] = r['evaluator_count']
        elif role == 'peer':
            by_person[key]['peer_score'] = float(r['avg_score'])
            by_person[key]['peer_count'] = r['evaluator_count']

    # 计算综合分（领导班子40% + 部门30% + 互评30%）
    results = []
    for p in by_person.values():
        ls, ds, ps = p['leader_score'], p['dept_score'], p['peer_score']
        # 综合分计算：能拿到几分就用几分，未填则按0计算（允许部分打分）
        # 标准是40+30+30=100%权重，但允许缺失
        total_w = 0
        total_v = 0
        if ls is not None:
            total_v += ls * 0.4
            total_w += 0.4
        if ds is not None:
            total_v += ds * 0.3
            total_w += 0.3
        if ps is not None:
            total_v += ps * 0.3
            total_w += 0.3
        final_score = round(total_v / total_w, 2) if total_w > 0 else None

        # 评级
        if final_score is None:
            level = '未完成'
        elif final_score >= 90:
            level = '优秀'
        elif final_score >= 80:
            level = '良好'
        elif final_score >= 60:
            level = '称职'
        else:
            level = '不称职'

        p['final_score'] = final_score
        p['level'] = level
        results.append(p)

    # 按综合分降序
    results.sort(key=lambda x: (x['final_score'] or 0), reverse=True)

    # 统计
    stats = {
        'total': len(results),
        'excellent': sum(1 for r in results if r['level'] == '优秀'),
        'good': sum(1 for r in results if r['level'] == '良好'),
        'competent': sum(1 for r in results if r['level'] == '称职'),
        'incompetent': sum(1 for r in results if r['level'] == '不称职'),
        'pending': sum(1 for r in results if r['level'] == '未完成'),
    }

    cur.close()
    conn.close()
    return jsonify({
        'cycle': cycle,
        'results': results,
        'stats': stats
    })


@bp.route('/api/evaluation/export', methods=['GET'])
def export_excel():
    """导出考核结果Excel（仅超管）"""
    user_phone = request.headers.get('X-User-Phone', '')
    if user_phone != '18184005669':
        return jsonify({'error': '无权限'}), 403

    cycle_id = request.args.get('cycle_id', type=int)
    conn = get_db()
    cur = conn.cursor()

    if cycle_id:
        cur.execute("SELECT * FROM evaluation_cycles WHERE id=%s", (cycle_id,))
    else:
        cur.execute("SELECT * FROM evaluation_cycles ORDER BY id DESC LIMIT 1")
    cycle = cur.fetchone()
    if not cycle:
        cur.close()
        conn.close()
        return jsonify({'error': '还没有考核周期'}), 404

    # 汇总打分
    cur.execute("""
        SELECT
            evaluatee_id, evaluatee_name, evaluatee_dept, evaluatee_project,
            evaluator_role,
            COUNT(*) as evaluator_count,
            ROUND(AVG(total_score)::numeric, 2) as avg_score
        FROM evaluation_scores
        WHERE cycle_id = %s
        GROUP BY evaluatee_id, evaluatee_name, evaluatee_dept, evaluatee_project, evaluator_role
    """, (cycle['id'],))
    grouped = cur.fetchall()

    by_person = {}
    for r in grouped:
        key = r['evaluatee_id']
        if key not in by_person:
            by_person[key] = {
                'id': r['evaluatee_id'],
                'name': r['evaluatee_name'],
                'dept': r['evaluatee_dept'] or '',
                'project': r['evaluatee_project'] or '',
                'leader_score': None, 'leader_count': 0,
                'dept_score': None, 'dept_count': 0,
                'peer_score': None, 'peer_count': 0,
            }
        role = r['evaluator_role']
        if role == 'leader':
            by_person[key]['leader_score'] = float(r['avg_score'])
            by_person[key]['leader_count'] = r['evaluator_count']
        elif role == 'dept_leader':
            by_person[key]['dept_score'] = float(r['avg_score'])
            by_person[key]['dept_count'] = r['evaluator_count']
        elif role == 'peer':
            by_person[key]['peer_score'] = float(r['avg_score'])
            by_person[key]['peer_count'] = r['evaluator_count']

    # 计算综合分
    rows = []
    for p in by_person.values():
        ls, ds, ps = p['leader_score'], p['dept_score'], p['peer_score']
        total_w, total_v = 0, 0
        if ls is not None: total_v += ls * 0.4; total_w += 0.4
        if ds is not None: total_v += ds * 0.3; total_w += 0.3
        if ps is not None: total_v += ps * 0.3; total_w += 0.3
        final_score = round(total_v / total_w, 2) if total_w > 0 else None
        if final_score is None: level = '未完成'
        elif final_score >= 90: level = '优秀'
        elif final_score >= 80: level = '良好'
        elif final_score >= 60: level = '称职'
        else: level = '不称职'
        rows.append({
            'id': p['id'],
            'name': p['name'],
            'dept': p['dept'],
            'project': p['project'],
            'leader_score': ls,
            'leader_count': p['leader_count'],
            'dept_score': ds,
            'dept_count': p['dept_count'],
            'peer_score': ps,
            'peer_count': p['peer_count'],
            'final_score': final_score,
            'level': level,
        })

    rows.sort(key=lambda x: (x['final_score'] or 0), reverse=True)

    # 生成 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "考核汇总"

    # 标题
    ws.merge_cells('A1:L1')
    ws['A1'] = f"{cycle['name']} 考核结果汇总"
    ws['A1'].font = Font(name='宋体', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 表头
    headers = ['序号', '姓名', '部门', '项目',
               '领导班子均分(40%)', '领导班子人数',
               '部门负责人均分(30%)', '部门人数',
               '员工互评均分(30%)', '互评人数',
               '综合分', '考核结果']
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.font = Font(name='宋体', size=11, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill('solid', fgColor='D9E1F2')
        c.border = Border(left=Side('thin'), right=Side('thin'),
                          top=Side('thin'), bottom=Side('thin'))
    ws.row_dimensions[2].height = 28

    # 数据
    for i, r in enumerate(rows, 1):
        row_data = [
            i,
            r['name'],
            r['dept'],
            r['project'],
            r['leader_score'] if r['leader_score'] is not None else '/',
            r['leader_count'],
            r['dept_score'] if r['dept_score'] is not None else '/',
            r['dept_count'],
            r['peer_score'] if r['peer_score'] is not None else '/',
            r['peer_count'],
            r['final_score'] if r['final_score'] is not None else '/',
            r['level'],
        ]
        for col, v in enumerate(row_data, 1):
            c = ws.cell(2 + i, col, v)
            c.font = Font(name='宋体', size=11)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = Border(left=Side('thin'), right=Side('thin'),
                              top=Side('thin'), bottom=Side('thin'))
            # 评级染色
            if col == 12:
                if v == '优秀': c.fill = PatternFill('solid', fgColor='C6EFCE')
                elif v == '良好': c.fill = PatternFill('solid', fgColor='FFEB9C')
                elif v == '不称职': c.fill = PatternFill('solid', fgColor='FFC7CE')

    # 列宽
    widths = [6, 12, 16, 16, 16, 14, 16, 12, 14, 12, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 统计sheet
    ws2 = wb.create_sheet('评分明细')
    ws2.append(['被考核人', '打分人', '打分人角色', '工作业绩(40)', '工作能力(20)',
                '工作态度(20)', '执行力(10)', '劳动纪律(10)', '总分', '评语', '提交时间'])
    for col in range(1, 12):
        c = ws2.cell(1, col)
        c.font = Font(name='宋体', size=11, bold=True)
        c.alignment = Alignment(horizontal='center')
        c.fill = PatternFill('solid', fgColor='D9E1F2')
        c.border = Border(left=Side('thin'), right=Side('thin'),
                          top=Side('thin'), bottom=Side('thin'))

    cur.execute("""
        SELECT evaluatee_name, evaluator_name, evaluator_role,
               performance_score, ability_score, attitude_score,
               execution_score, discipline_score, total_score, comment, submitted_at
        FROM evaluation_scores
        WHERE cycle_id = %s
        ORDER BY evaluatee_name, submitted_at
    """, (cycle['id'],))
    detail_rows = cur.fetchall()
    role_map = {'leader': '领导班子', 'dept_leader': '部门负责人', 'peer': '员工互评'}
    for r in detail_rows:
        ws2.append([
            r['evaluatee_name'], r['evaluator_name'],
            role_map.get(r['evaluator_role'], r['evaluator_role']),
            float(r['performance_score']), float(r['ability_score']),
            float(r['attitude_score']), float(r['execution_score']),
            float(r['discipline_score']), float(r['total_score']),
            r['comment'] or '',
            r['submitted_at'].strftime('%Y-%m-%d %H:%M') if r['submitted_at'] else ''
        ])
    widths2 = [12, 12, 12, 14, 12, 12, 10, 12, 8, 30, 16]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    cur.close()
    conn.close()

    # 输出
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    from flask import send_file
    filename = f"{cycle['period']}_evaluation_results.xlsx"
    return send_file(
        bio,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

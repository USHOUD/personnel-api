#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""安装公司人员信息管理系统 - Supabase版"""

from flask import Flask, request, jsonify
from flask_cors import CORS
from datetime import datetime, timedelta
import hashlib
from supabase_config import get_db, init_db
from psycopg2.extras import RealDictCursor

app = Flask(__name__)
CORS(app, expose_headers=['X-User-Phone'])

# 初始化数据库
init_db()

# ============= 权限管理 =============

def get_current_user():
    """从请求头获取当前用户信息"""
    phone = request.headers.get('X-User-Phone', '')
    if not phone:
        return None
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT phone, name, is_admin FROM users WHERE phone=%s", (phone,))
    user = cur.fetchone()
    cur.close()
    conn.close()
    return user

def mask_person_data(p, is_self=False, is_admin=False):
    """根据权限过滤人员数据"""
    dept = p.get('dept', '') or ''
    if not dept:
        dept = get_dept(p)
    
    # 管理员或本人看完整信息
    if is_admin or is_self:
        return {
            'id': p['id'],
            'name': p['name'],
            'gender': p.get('gender', '') or '',
            'id_card': p.get('id_card', '') or '',
            'birth': p.get('birth', '') or '',
            'edu': p.get('edu', '') or '',
            'hometown': p.get('hometown', '') or '',
            'position': p.get('position', '') or '',
            'dept': dept,
            'project': p.get('project', '') or '未分配',
            'phone': p.get('phone', '') or '',
            'cert': p.get('cert', '') or '',
            'category': p.get('category', ''),
            'salary': float(p['salary']) if p.get('salary') else None,
            'status': p.get('status', '') or '在岗',
            'status_detail': p.get('status_detail', '') or '',
            'hire_date': p.get('hire_date', '') or '',
            'leave_date': p.get('leave_date', '') or ''
        }
    
    # 普通用户看别人：只显示基本信息
    return {
        'id': p['id'],
        'name': p['name'],
        'gender': p.get('gender', '') or '',
        'position': p.get('position', '') or '',
        'dept': dept,
        'project': p.get('project', '') or '未分配',
        'phone': p.get('phone', '') or '',
        'category': p.get('category', ''),
        'status': p.get('status', '') or '在岗'
    }

# ============= 人员相关API =============

def get_dept(person):
    """根据职位和项目自动推断部门"""
    position = (person.get('position', '') or '').lower()
    project = person.get('project', '') or ''
    
    # 领导班子
    if any(k in position for k in ['经理', '书记']):
        return '领导班子'
    
    # 后台人员
    if project == '后台':
        if any(k in position for k in ['bim', '设计']):
            return '质量技术部'
        elif any(k in position for k in ['策划', '调度']):
            return '生产管理中心'
        elif any(k in position for k in ['商务', '预算', '造价', '结算', '成本', '核算']):
            return '商务法务部'
        elif any(k in position for k in ['财务', '会计']):
            return '财务部'
        elif any(k in position for k in ['安全']):
            return '安全环保部'
        else:
            return '综合办公室'
    
    # 项目人员
    if any(k in position for k in ['商务', '预算', '造价', '结算', '成本', '核算']):
        return '商务法务部'
    elif any(k in position for k in ['安全']):
        return '安全环保部'
    else:
        return '工程技术部'

def auto_restore_on_duty():
    """自动恢复在岗：检查出差/休假是否过期，过期则恢复为在岗"""
    try:
        import re
        # 北京时间 = UTC+8
        now_bj = datetime.utcnow() + timedelta(hours=8)
        today_str = now_bj.strftime('%Y-%m-%d')

        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id, status, status_detail FROM personnel WHERE status IN ('出差','休假')")
        rows = cur.fetchall()

        restored = []
        for row in rows:
            detail = row.get('status_detail', '') or ''
            # 提取日期：匹配 至YYYY-MM-DD 或 (至YYYY-MM-DD)
            m = re.search(r'至(\d{4}-\d{2}-\d{2})', detail)
            if m:
                end_date = m.group(1)
                if today_str > end_date:
                    cur.execute("UPDATE personnel SET status='在岗', status_detail='', updated_at=NOW() WHERE id=%s", (row['id'],))
                    restored.append(row['id'])

        if restored:
            conn.commit()
        cur.close()
        conn.close()
        return restored
    except Exception as e:
        print(f"auto_restore_on_duty error: {e}")
        return []

@app.route('/api/personnel')
def get_personnel():
    """获取人员列表"""
    # 自动恢复过期的出差/休假人员
    auto_restore_on_duty()

    conn = get_db()
    cur = conn.cursor()
    
    category = request.args.get('category', 'all')
    search = request.args.get('search', '').strip()
    
    if category == 'formal':
        cur.execute("SELECT * FROM personnel WHERE category='正式职工'")
    elif category == 'outsourced':
        cur.execute("SELECT * FROM personnel WHERE category IN ('C1','C2')")
    elif category == 'C1':
        cur.execute("SELECT * FROM personnel WHERE category='C1'")
    elif category == 'C2':
        cur.execute("SELECT * FROM personnel WHERE category='C2'")
    else:
        cur.execute("SELECT * FROM personnel")
    
    people = cur.fetchall()
    
    if search:
        key = search.lower()
        people = [p for p in people if key in (p['name'] or '').lower() 
                  or key in (p['position'] or '').lower()
                  or key in (p['project'] or '').lower()]
    
    # 排序逻辑
    def sort_key(p):
        # 1. 固定人员优先
        fixed_order = {'邱方恒': 0, '廖志成': 1, '吕亮': 2, '李强': 3}
        fixed = fixed_order.get(p['name'], 99)
        
        # 2. 人员类别
        cat_order = {'正式职工': 0, 'C1': 1, 'C2': 2}
        cat = cat_order.get(p.get('category', ''), 3)
        
        # 3. 项目优先级
        project = p.get('project', '') or ''
        if project == '后台':
            proj = 0
        elif project and project != '其他':
            proj = 1
        else:
            proj = 2
        
        # 4. 职务级别
        position = (p.get('position', '') or '').lower()
        if any(k in position for k in ['经理', '书记']):
            pos = 0
        elif any(k in position for k in ['部长', '主管', '副部长']):
            pos = 1
        else:
            pos = 2
        
        return (fixed, cat, proj, pos, p.get('name', ''))
    
    people.sort(key=sort_key)
    
    # 获取当前用户权限
    current_user = get_current_user()
    is_admin = current_user['is_admin'] if current_user else False
    current_user_phone = current_user['phone'] if current_user else ''
    
    # 找到当前用户的personnel_id
    current_person_id = None
    if current_user_phone:
        conn2 = get_db()
        cur2 = conn2.cursor()
        cur2.execute("SELECT id FROM personnel WHERE phone=%s", (current_user_phone,))
        row = cur2.fetchone()
        if row:
            current_person_id = row['id']
        cur2.close()
        conn2.close()
    
    # 转换为JSON格式
    result = []
    for p in people:
        is_self = (p['id'] == current_person_id)
        result.append(mask_person_data(p, is_self=is_self, is_admin=is_admin))
    
    cur.close()
    conn.close()
    return jsonify(result)

@app.route('/api/personnel/<person_id>')
def get_person(person_id):
    """获取单个人员详情"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM personnel WHERE id=%s", (person_id,))
    p = cur.fetchone()
    cur.close()
    conn.close()
    
    if not p:
        return jsonify({'error': '未找到该人员'}), 404
    
    # 权限检查
    current_user = get_current_user()
    is_admin = current_user['is_admin'] if current_user else False
    current_user_phone = current_user['phone'] if current_user else ''
    
    is_self = (p.get('phone', '') == current_user_phone)
    
    return jsonify(mask_person_data(p, is_self=is_self, is_admin=is_admin))

@app.route('/api/personnel', methods=['POST'])
def add_person():
    """新增人员"""
    data = request.json
    if not data.get('name'):
        return jsonify({'error': '姓名不能为空'}), 400
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # 生成ID（找最大编号+1，避免重复）
        cat = data.get('category', '正式职工')
        if cat == '正式职工':
            cur.execute("SELECT id FROM personnel WHERE id LIKE 'F%' ORDER BY CAST(SUBSTRING(id FROM 2) AS INTEGER) DESC LIMIT 1")
        else:
            cur.execute("SELECT id FROM personnel WHERE id LIKE 'O%' ORDER BY CAST(SUBSTRING(id FROM 2) AS INTEGER) DESC LIMIT 1")
        
        row = cur.fetchone()
        if row and row['id']:
            max_num = int(row['id'][1:])
            pid = f"{'F' if cat == '正式职工' else 'O'}{max_num + 1}"
        else:
            pid = f"{'F' if cat == '正式职工' else 'O'}1"
        
        # salary转数字，空字符串转None
        salary = data.get('salary')
        if salary == '' or salary is None:
            salary = None
        else:
            try:
                salary = float(salary)
            except:
                salary = None
        
        cur.execute("""
            INSERT INTO personnel (id, name, gender, id_card, birth, edu, hometown, position, dept, project, phone, cert, category, salary, hire_date)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (pid, data['name'], data.get('gender',''), data.get('id_card',''),
              data.get('birth',''), data.get('edu',''), data.get('hometown',''),
              data.get('position',''), data.get('dept',''),
              data.get('project','未分配'),
              data.get('phone',''), data.get('cert',''), cat, salary,
              data.get('hire_date', '') or None))
        
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'success': True, 'person': {'id': pid, **data}})
    except Exception as e:
        print(f"add_person error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/personnel/<person_id>', methods=['PUT'])
def update_person(person_id):
    """更新人员信息"""
    data = request.json
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute("SELECT id FROM personnel WHERE id=%s", (person_id,))
        if not cur.fetchone():
            cur.close(); conn.close()
            return jsonify({'error': '未找到该人员'}), 404
        
        fields = []
        values = []
        for key in ['name','gender','id_card','birth','edu','hometown','position','dept','project','phone','cert','category','salary','status','status_detail','hire_date','leave_date']:
            if key in data:
                val = data[key]
                # salary空字符串转None
                if key == 'salary' and (val == '' or val is None):
                    val = None
                fields.append(f"{key}=%s")
                values.append(val)
        
        if fields:
            fields.append("updated_at=NOW()")
            values.append(person_id)
            cur.execute(f"UPDATE personnel SET {','.join(fields)} WHERE id=%s", values)
            conn.commit()
        
        cur.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        print(f"update_person error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/personnel/<person_id>', methods=['DELETE'])
def delete_person(person_id):
    """删除人员"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM personnel WHERE id=%s", (person_id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'success': True})

# ============= 状态管理API =============

@app.route('/api/personnel/<person_id>/status', methods=['PUT'])
def update_status(person_id):
    """更新人员状态"""
    data = request.json
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE personnel SET status=%s, status_detail=%s, updated_at=NOW() WHERE id=%s",
                (data.get('status','在岗'), data.get('detail',''), person_id))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/personnel/<person_id>/return', methods=['PUT'])
def person_return(person_id):
    """归队/销假"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE personnel SET status='在岗', status_detail='', updated_at=NOW() WHERE id=%s", (person_id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/personnel/<person_id>/resign', methods=['PUT'])
def person_resign(person_id):
    """离职：设置leave_date和状态为已离职"""
    try:
        data = request.json or {}
        conn = get_db()
        cur = conn.cursor()

        cur.execute("SELECT id FROM personnel WHERE id=%s", (person_id,))
        if not cur.fetchone():
            cur.close(); conn.close()
            return jsonify({'error': '未找到该人员'}), 404

        leave_date = data.get('leave_date', '') or datetime.utcnow().strftime('%Y-%m-%d')
        cur.execute("UPDATE personnel SET status='已离职', leave_date=%s, status_detail='', updated_at=NOW() WHERE id=%s",
                    (leave_date, person_id))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        print(f"person_resign error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/personnel/<person_id>/hire', methods=['PUT'])
def person_hire(person_id):
    """入职/重新入职：设置hire_date和状态为在岗"""
    try:
        data = request.json or {}
        conn = get_db()
        cur = conn.cursor()

        cur.execute("SELECT id FROM personnel WHERE id=%s", (person_id,))
        if not cur.fetchone():
            cur.close(); conn.close()
            return jsonify({'error': '未找到该人员'}), 404

        hire_date = data.get('hire_date', '') or datetime.utcnow().strftime('%Y-%m-%d')
        cur.execute("UPDATE personnel SET status='在岗', hire_date=%s, leave_date=NULL, status_detail='', updated_at=NOW() WHERE id=%s",
                    (hire_date, person_id))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        print(f"person_hire error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/auto-restore', methods=['POST'])
def auto_restore_api():
    """手动触发自动恢复在岗（也供前端定期调用）"""
    restored = auto_restore_on_duty()
    return jsonify({'success': True, 'restored': restored, 'count': len(restored)})

# ============= 假期余额API =============

@app.route('/api/leave-balance/<person_id>')
def get_leave_balance(person_id):
    """获取人员假期余额"""
    try:
        import re
        from datetime import datetime as dt
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("SELECT * FROM personnel WHERE id=%s", (person_id,))
        person = cur.fetchone()
        if not person:
            cur.close(); conn.close()
            return jsonify({'error': '未找到该人员'}), 404
        
        # 权限检查：仅本人或管理员可查看假期余额
        current_user = get_current_user()
        is_admin = current_user and current_user.get('is_admin', False)
        is_self = current_user and current_user.get('phone') == person.get('phone')
        if not is_admin and not is_self:
            cur.close(); conn.close()
            return jsonify({'error': '无权限查看'}), 403
        
        category = person.get('category', '')
        hire_date = person.get('hire_date', '') or ''
        birth = person.get('birth', '') or ''
        
        now = dt.now()
        current_year = now.year
        
        # 计算工龄（正式职工用参加工作时间，外包用入职时间）
        work_years = 0
        if category == '正式职工':
            # 正式职工从出生+22岁估算，或从hire_date
            if hire_date:
                try:
                    hd = dt.strptime(hire_date[:10], '%Y-%m-%d')
                    work_years = (now - hd).days / 365.25
                except:
                    work_years = 0
        else:
            # 外包人员从hire_date算
            if hire_date:
                try:
                    hd = dt.strptime(hire_date[:10], '%Y-%m-%d')
                    work_years = (now - hd).days / 365.25
                except:
                    work_years = 0
        
        # 年休假天数
        annual_leave = 0
        if category == '正式职工':
            if work_years >= 20:
                annual_leave = 15
            elif work_years >= 10:
                annual_leave = 10
            elif work_years >= 1:
                annual_leave = 5
        else:
            # 外包：满1年后5天
            if work_years >= 1:
                annual_leave = 5
        
        # 探亲假（正式职工，工作满1年）
        family_leave = 0
        if category == '正式职工' and work_years >= 1:
            family_leave = 30  # 默认探配偶30天
        
        # 统计当年已使用假期
        cur.execute("""
            SELECT reason, start_date, end_date FROM leave_records 
            WHERE person_id=%s AND status='已通过' 
            AND start_date >= %s AND start_date < %s
        """, (person_id, f'{current_year}-01-01', f'{current_year + 1}-01-01'))
        used_records = cur.fetchall()
        
        used_annual = 0
        used_family = 0
        used_other = 0
        used_details = []
        
        for r in used_records:
            try:
                start = dt.strptime(r['start_date'], '%Y-%m-%d')
                end = dt.strptime(r['end_date'], '%Y-%m-%d')
                days = (end - start).days + 1
            except:
                days = 0
            
            reason = r.get('reason', '') or ''
            used_details.append({'reason': reason, 'days': days, 'start': r['start_date'], 'end': r['end_date']})
            
            if '年休' in reason:
                used_annual += days
            elif '探亲' in reason:
                used_family += days
            else:
                used_other += days
        
        cur.close()
        conn.close()
        
        # 从考勤表统计（更准确）
        # 先用leave_records的统计
        
        result = {
            'person_id': person_id,
            'name': person['name'],
            'category': category,
            'hire_date': hire_date,
            'work_years': round(work_years, 1),
            'leave_types': []
        }
        
        if annual_leave > 0:
            result['leave_types'].append({
                'type': '年休假',
                'total': annual_leave,
                'used': used_annual,
                'remaining': max(0, annual_leave - used_annual)
            })
        
        if category == '正式职工' and family_leave > 0:
            family_used_flag = 1 if used_family > 0 else 0
            result['leave_types'].append({
                'type': '探亲假（探配偶）',
                'total': f'{family_leave}天/年（一次）',
                'used': f'已休{used_family}天' if used_family > 0 else '未休',
                'remaining': 0 if used_family > 0 else family_leave
            })
        
        # 其他假期类型（不扣额度）
        if used_other > 0:
            result['leave_types'].append({
                'type': '其他假（病/陪产/事假等）',
                'total': '-',
                'used': used_other,
                'remaining': '-'
            })
        
        result['used_details'] = used_details
        
        return jsonify(result)
    except Exception as e:
        print(f"leave_balance error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/leave-balance')
def get_all_leave_balance():
    """获取所有人员假期余额"""
    try:
        from datetime import datetime as dt
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("SELECT * FROM personnel WHERE status != '已离职' ORDER BY category, name")
        people = cur.fetchall()
        
        now = dt.now()
        current_year = now.year
        
        # 获取当年所有已审批休假
        cur.execute("""
            SELECT person_id, reason, start_date, end_date FROM leave_records 
            WHERE status='已通过' AND start_date >= %s AND start_date < %s
        """, (f'{current_year}-01-01', f'{current_year + 1}-01-01'))
        all_used = cur.fetchall()
        
        # 按人员分组
        used_map = {}
        for r in all_used:
            pid = r['person_id']
            if pid not in used_map:
                used_map[pid] = []
            used_map[pid].append(r)
        
        results = []
        for p in people:
            category = p.get('category', '')
            hire_date = p.get('hire_date', '') or ''
            
            work_years = 0
            if hire_date:
                try:
                    hd = dt.strptime(hire_date[:10], '%Y-%m-%d')
                    work_years = (now - hd).days / 365.25
                except:
                    pass
            
            annual_leave = 0
            if category == '正式职工':
                if work_years >= 20: annual_leave = 15
                elif work_years >= 10: annual_leave = 10
                elif work_years >= 1: annual_leave = 5
            else:
                if work_years >= 1: annual_leave = 5
            
            used_annual = 0
            used_family = 0
            for r in used_map.get(p['id'], []):
                try:
                    s = dt.strptime(r['start_date'], '%Y-%m-%d')
                    e = dt.strptime(r['end_date'], '%Y-%m-%d')
                    days = (e - s).days + 1
                except:
                    days = 0
                reason = r.get('reason', '') or ''
                if '年休' in reason: used_annual += days
                elif '探亲' in reason: used_family += days
            
            entry = {
                'id': p['id'],
                'name': p['name'],
                'category': category,
                'work_years': round(work_years, 1),
                'annual_total': annual_leave,
                'annual_used': used_annual,
                'annual_remaining': max(0, annual_leave - used_annual),
            }
            
            if category == '正式职工':
                entry['family_total'] = 30 if work_years >= 1 else 0
                entry['family_used'] = used_family
                entry['family_remaining'] = 0 if used_family > 0 else entry['family_total']
            
            results.append(entry)
        
        cur.close()
        conn.close()
        return jsonify(results)
    except Exception as e:
        print(f"all_leave_balance error: {e}")
        return jsonify({'error': str(e)}), 500

# ============= 调动记录API =============

@app.route('/api/transfers')
def get_transfers():
    """获取调动记录"""
    conn = get_db()
    cur = conn.cursor()
    person_id = request.args.get('person_id')
    
    if person_id:
        cur.execute("SELECT * FROM transfers WHERE person_id=%s ORDER BY created_at DESC", (person_id,))
    else:
        cur.execute("SELECT * FROM transfers ORDER BY created_at DESC")
    
    records = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify([dict(r) for r in records])

@app.route('/api/transfers', methods=['POST'])
def add_transfer():
    """新增调动记录"""
    data = request.json
    if not data.get('person_id') or not data.get('to_project'):
        return jsonify({'error': '信息不完整'}), 400
    
    conn = get_db()
    cur = conn.cursor()
    
    # 获取原项目
    cur.execute("SELECT project FROM personnel WHERE id=%s", (data['person_id'],))
    person = cur.fetchone()
    from_project = person['project'] if person else ''
    
    # 生成ID
    cur.execute("SELECT COUNT(*) FROM transfers")
    tid = f"T{cur.fetchone()['count'] + 1}"
    
    cur.execute("""
        INSERT INTO transfers (id, person_id, person_name, from_project, to_project, transfer_date, notes, created_at)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
    """, (tid, data['person_id'], data.get('person_name',''), from_project,
          data['to_project'], data.get('transfer_date',''), data.get('notes',''),
          datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    
    # 更新人员项目
    cur.execute("UPDATE personnel SET project=%s, updated_at=NOW() WHERE id=%s",
                (data['to_project'], data['person_id']))
    
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'success': True, 'transfer': {'id': tid, **data}})

@app.route('/api/personnel/<person_id>/timeline')
def get_timeline(person_id):
    """获取项目时间线"""
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("SELECT * FROM transfers WHERE person_id=%s ORDER BY transfer_date", (person_id,))
    transfers = cur.fetchall()
    
    cur.execute("SELECT project FROM personnel WHERE id=%s", (person_id,))
    person = cur.fetchone()
    current_project = person['project'] if person else '未分配'
    
    timeline = []
    if not transfers:
        timeline.append({'project': current_project, 'start_date': '至今', 'end_date': '至今', 'months': 12})
    
    cur.close()
    conn.close()
    return jsonify({
        'person_id': person_id,
        'current_project': current_project,
        'timeline': timeline,
        'transfers': [dict(t) for t in transfers]
    })

# ============= 休假申请API =============

@app.route('/api/leave')
def get_leave():
    """获取休假申请"""
    conn = get_db()
    cur = conn.cursor()
    person_id = request.args.get('person_id')
    
    if person_id:
        cur.execute("SELECT * FROM leave_records WHERE person_id=%s ORDER BY created_at DESC", (person_id,))
    else:
        cur.execute("SELECT * FROM leave_records ORDER BY created_at DESC")
    
    records = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify([dict(r) for r in records])

@app.route('/api/leave', methods=['POST'])
def add_leave():
    """新增休假申请"""
    data = request.json
    if not data.get('person_id') or not data.get('start_date') or not data.get('end_date'):
        return jsonify({'error': '信息不完整'}), 400
    
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("SELECT COUNT(*) FROM leave_records")
    lid = f"L{cur.fetchone()['count'] + 1}"
    
    cur.execute("""
        INSERT INTO leave_records (id, person_id, person_name, start_date, end_date, reason, status, created_at)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
    """, (lid, data['person_id'], data.get('person_name',''),
          data['start_date'], data['end_date'], data.get('reason',''),
          '待审批', datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'success': True, 'record': {'id': lid, **data}})

@app.route('/api/leave/<record_id>/approve', methods=['PUT'])
def approve_leave(record_id):
    """审批休假"""
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("UPDATE leave_records SET status='已通过' WHERE id=%s RETURNING person_id, end_date", (record_id,))
    record = cur.fetchone()
    
    if record:
        cur.execute("UPDATE personnel SET status='休假', status_detail=%s, updated_at=NOW() WHERE id=%s",
                    (f"休假至{record['end_date']}", record['person_id']))
        conn.commit()
    
    cur.close()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/leave/<record_id>/reject', methods=['PUT'])
def reject_leave(record_id):
    """驳回休假"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE leave_records SET status='已驳回' WHERE id=%s", (record_id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'success': True})

# ============= 出差单API =============

@app.route('/api/trip')
def get_trips():
    """获取出差单"""
    conn = get_db()
    cur = conn.cursor()
    person_id = request.args.get('person_id')
    
    if person_id:
        cur.execute("SELECT * FROM trip_records WHERE person_id=%s ORDER BY created_at DESC", (person_id,))
    else:
        cur.execute("SELECT * FROM trip_records ORDER BY created_at DESC")
    
    records = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify([dict(r) for r in records])

@app.route('/api/trip', methods=['POST'])
def add_trip():
    """新增出差单"""
    data = request.json
    if not data.get('person_id') or not data.get('destination'):
        return jsonify({'error': '信息不完整'}), 400
    
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("SELECT COUNT(*) FROM trip_records")
    bid = f"B{cur.fetchone()['count'] + 1}"
    
    cur.execute("""
        INSERT INTO trip_records (id, person_id, person_name, destination, start_date, end_date, reason, status, created_at)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (bid, data['person_id'], data.get('person_name',''),
          data['destination'], data.get('start_date',''), data.get('end_date',''),
          data.get('reason',''), '待审批', datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'success': True, 'record': {'id': bid, **data}})

@app.route('/api/trip/<record_id>/approve', methods=['PUT'])
def approve_trip(record_id):
    """审批出差"""
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("UPDATE trip_records SET status='已通过' WHERE id=%s RETURNING person_id, destination", (record_id,))
    record = cur.fetchone()
    
    if record:
        # 获取出差结束日期用于自动恢复
        cur.execute("SELECT end_date FROM trip_records WHERE id=%s", (record_id,))
        trip = cur.fetchone()
        end_str = ''
        if trip and trip.get('end_date'):
            end_str = f"(至{trip['end_date']})"
        cur.execute("UPDATE personnel SET status='出差', status_detail=%s, updated_at=NOW() WHERE id=%s",
                    (f"出差-{record['destination']}{end_str}", record['person_id']))
        conn.commit()
    
    cur.close()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/trip/<record_id>/reject', methods=['PUT'])
def reject_trip(record_id):
    """驳回出差"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE trip_records SET status='已驳回' WHERE id=%s", (record_id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'success': True})

# ============= 统计API =============

@app.route('/api/statistics')
def get_statistics():
    """获取统计数据"""
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("SELECT COUNT(*) as total FROM personnel")
    total = cur.fetchone()['total']
    
    cur.execute("SELECT COUNT(*) as c FROM personnel WHERE category='正式职工'")
    formal = cur.fetchone()['c']
    
    cur.execute("SELECT COUNT(*) as c FROM personnel WHERE category='C1'")
    c1 = cur.fetchone()['c']
    
    cur.execute("SELECT COUNT(*) as c FROM personnel WHERE category='C2'")
    c2 = cur.fetchone()['c']
    
    cur.execute("SELECT COUNT(*) as c FROM personnel WHERE gender='男'")
    male = cur.fetchone()['c']
    
    cur.execute("SELECT COUNT(*) as c FROM personnel WHERE gender='女'")
    female = cur.fetchone()['c']
    
    # 学历统计
    def map_edu(edu):
        if not edu: return '未知'
        if any(k in edu for k in ['硕士','研究生']): return '硕士研究生'
        if any(k in edu for k in ['本科','大学']): return '本科'
        if any(k in edu for k in ['专科','大专']): return '专科'
        if any(k in edu for k in ['中专','初中','高中']): return '高中及以下'
        return '未知'
    
    cur.execute("SELECT edu FROM personnel")
    edu_stats = {}
    for row in cur.fetchall():
        cat = map_edu(row['edu'])
        edu_stats[cat] = edu_stats.get(cat, 0) + 1
    
    # 项目统计
    cur.execute("SELECT project, COUNT(*) as c FROM personnel WHERE category='正式职工' GROUP BY project")
    dept_stats = {row['project']: row['c'] for row in cur.fetchall()}
    
    # 证书统计
    cur.execute("SELECT name, cert FROM personnel")
    cert_stats = {
        '一建': {'count': 0, 'persons': []},
        '一造': {'count': 0, 'persons': []},
        '二建': {'count': 0, 'persons': []},
        '二造': {'count': 0, 'persons': []},
        '八大员': {'count': 0, 'detail': {}},
        '其他': {'count': 0, 'persons': []},
        '无证书': {'count': 0, 'persons': []}
    }
    total_with_cert = 0
    
    # 八大员证书关键词分类
    badayuan_types = ['质量员', '施工员', '安全员', '测量员', '资料员', '材料员', '机械员', '劳务员', '标准员', '试验员']
    
    for row in cur.fetchall():
        name = row['name']
        cert = row['cert']
        if cert and cert.strip() and cert != '/':
            total_with_cert += 1
            if '一建' in cert or '一级建造师' in cert:
                cert_stats['一建']['count'] += 1
                cert_stats['一建']['persons'].append(name)
            elif '一造' in cert or '一级造价' in cert:
                cert_stats['一造']['count'] += 1
                cert_stats['一造']['persons'].append(name)
            elif '二建' in cert or '二级建造师' in cert:
                cert_stats['二建']['count'] += 1
                cert_stats['二建']['persons'].append(name)
            elif '二造' in cert or '二级造价' in cert:
                cert_stats['二造']['count'] += 1
                cert_stats['二造']['persons'].append(name)
            elif any(k in cert for k in badayuan_types):
                cert_stats['八大员']['count'] += 1
                # 按证书子类型分组
                found_type = '其他'
                for t in badayuan_types:
                    if t in cert:
                        found_type = t
                        break
                if found_type not in cert_stats['八大员']['detail']:
                    cert_stats['八大员']['detail'][found_type] = []
                cert_stats['八大员']['detail'][found_type].append(name)
            else:
                cert_stats['其他']['count'] += 1
                cert_stats['其他']['persons'].append(f"{name}（{cert}）")
        else:
            cert_stats['无证书']['count'] += 1
            cert_stats['无证书']['persons'].append(name)
    
    # 一建指标
    cur.execute("SELECT * FROM exam_targets WHERE exam_type='一建' LIMIT 1")
    exam_row = cur.fetchone()
    exam_target = None
    if exam_row:
        cur.execute("SELECT person_name FROM exam_target_persons WHERE target_id=%s ORDER BY id", (exam_row['id'],))
        persons = [r['person_name'] for r in cur.fetchall()]
        exam_target = {
            'id': exam_row['id'],
            'exam_type': exam_row['exam_type'],
            'year': exam_row['year'],
            'persons': persons
        }
    
    cur.close()
    conn.close()
    
    return jsonify({
        'total': total,
        'formal_count': formal,
        'outsourced_count': c1 + c2,
        'c1_count': c1,
        'c2_count': c2,
        'gender': {'male': male, 'female': female},
        'edu': edu_stats,
        'dept': dept_stats,
        'cert': cert_stats,
        'exam_target': exam_target
    })

# ============= 数据导出API =============

@app.route('/api/export')
def export_data():
    """导出人员数据为Excel"""
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from flask import send_file
        
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM personnel")
        people = cur.fetchall()
        cur.close()
        conn.close()
        
        # 排序（与小程序一致）
        def sort_key(p):
            fixed_order = {'邱方恒': 0, '廖志成': 1, '吕亮': 2, '李强': 3}
            fixed = fixed_order.get(p['name'], 99)
            cat_order = {'正式职工': 0, 'C1': 1, 'C2': 2}
            cat = cat_order.get(p.get('category', ''), 3)
            project = p.get('project', '') or ''
            if project == '后台': proj = 0
            elif project and project != '其他': proj = 1
            else: proj = 2
            position = (p.get('position', '') or '').lower()
            if any(k in position for k in ['经理', '书记']): pos = 0
            elif any(k in position for k in ['部长', '主管', '副部长']): pos = 1
            else: pos = 2
            return (fixed, cat, proj, pos, p.get('name', ''))
        
        people.sort(key=sort_key)
        
        wb = Workbook()
        ws = wb.active
        ws.title = '人员名单'
        
        # 表头
        headers = ['序号', 'ID', '姓名', '性别', '身份证号', '出生日期', '学历', '籍贯',
                   '岗位', '部门', '项目', '电话', '证书', '类别', '工资',
                   '状态', '状态详情', '入职日期', '离职日期']
        
        thin = Side(style='thin')
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_font = Font(name='宋体', size=11, bold=True)
        data_font = Font(name='宋体', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border_all
        
        # 数据
        for i, p in enumerate(people):
            dept = p.get('dept', '') or ''
            if not dept:
                dept = get_dept(p)
            
            row_data = [
                i + 1,
                p['id'],
                p['name'],
                p.get('gender', ''),
                p.get('id_card', ''),
                p.get('birth', ''),
                p.get('edu', ''),
                p.get('hometown', ''),
                p.get('position', ''),
                dept,
                p.get('project', ''),
                p.get('phone', ''),
                p.get('cert', ''),
                p.get('category', ''),
                float(p['salary']) if p.get('salary') else '',
                p.get('status', ''),
                p.get('status_detail', ''),
                p.get('hire_date', ''),
                p.get('leave_date', '') or ''
            ]
            
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i + 2, column=col, value=val)
                cell.font = data_font
                cell.alignment = center_align
                cell.border = border_all
        
        # 自动列宽
        for col in range(1, len(headers) + 1):
            max_len = len(str(headers[col - 1]))
            for row in range(2, len(people) + 2):
                val = ws.cell(row=row, column=col).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 4, 30)
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='安装公司人员名单.xlsx'
        )
    except Exception as e:
        print(f"export error: {e}")
        return jsonify({'error': str(e)}), 500

# ============= 登录API =============

# ============= 一建指标API =============

@app.route('/api/exam-targets', methods=['GET'])
def get_exam_targets():
    """获取一建指标"""
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("SELECT * FROM exam_targets ORDER BY year DESC LIMIT 1")
    target = cur.fetchone()
    
    if target:
        cur.execute("SELECT person_name FROM exam_target_persons WHERE target_id=%s ORDER BY id", (target['id'],))
        persons = [r['person_name'] for r in cur.fetchall()]
        result = {
            'id': target['id'],
            'exam_type': target['exam_type'],
            'year': target['year'],
            'persons': persons
        }
    else:
        result = None
    
    cur.close()
    conn.close()
    return jsonify(result)

@app.route('/api/exam-targets', methods=['POST'])
def save_exam_targets():
    """保存一建指标"""
    data = request.json
    exam_type = data.get('exam_type', '一建')
    year = data.get('year', 2026)
    persons = data.get('persons', [])
    
    conn = get_db()
    cur = conn.cursor()
    
    try:
        # 查找或创建指标
        cur.execute("SELECT id FROM exam_targets WHERE exam_type=%s AND year=%s", (exam_type, year))
        row = cur.fetchone()
        
        if row:
            target_id = row['id']
            # 更新时间
            cur.execute("UPDATE exam_targets SET updated_at=NOW() WHERE id=%s", (target_id,))
            # 删除旧人员
            cur.execute("DELETE FROM exam_target_persons WHERE target_id=%s", (target_id,))
        else:
            cur.execute("INSERT INTO exam_targets (exam_type, year) VALUES (%s, %s) RETURNING id", (exam_type, year))
            target_id = cur.fetchone()['id']
        
        # 插入新人员
        for name in persons:
            cur.execute("INSERT INTO exam_target_persons (target_id, person_name) VALUES (%s, %s)", (target_id, name))
        
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        cur.close()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/login', methods=['POST'])
def login():
    """用户登录"""
    data = request.json
    phone = data.get('phone','').strip()
    password = data.get('password','').strip()
    
    if not phone or not password:
        return jsonify({'error': '手机号和密码不能为空'}), 400
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE phone=%s", (phone,))
    user = cur.fetchone()
    cur.close()
    conn.close()
    
    if not user:
        return jsonify({'error': '用户不存在'}), 401
    if user['password'] != password:
        return jsonify({'error': '密码错误'}), 401
    
    token = hashlib.md5(f"{phone}:{password}:{datetime.now().date()}".encode()).hexdigest()
    
    return jsonify({
        'success': True,
        'token': token,
        'user': {
            'phone': user['phone'],
            'name': user['name'],
            'is_admin': user['is_admin']
        }
    })

@app.route('/api/change-password', methods=['PUT'])
def change_password():
    """修改密码"""
    data = request.json
    phone = data.get('phone', '').strip()
    old_password = data.get('old_password', '').strip()
    new_password = data.get('new_password', '').strip()
    
    if not phone or not old_password or not new_password:
        return jsonify({'error': '信息不完整'}), 400
    
    if len(new_password) < 6:
        return jsonify({'error': '新密码至少6位'}), 400
    
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT password FROM users WHERE phone=%s", (phone,))
        user = cur.fetchone()
        
        if not user:
            cur.close(); conn.close()
            return jsonify({'error': '用户不存在'}), 404
        
        if user['password'] != old_password:
            cur.close(); conn.close()
            return jsonify({'error': '原密码错误'}), 401
        
        cur.execute("UPDATE users SET password=%s WHERE phone=%s", (new_password, phone))
        conn.commit()
        cur.close(); conn.close()
        return jsonify({'success': True, 'message': '密码修改成功'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reset-password', methods=['POST'])
def reset_password():
    """重置密码（仅超级管理员18184005669）"""
    data = request.json
    phone = data.get('phone', '').strip()
    admin_phone = data.get('admin_phone', '').strip()
    
    if not phone:
        return jsonify({'error': '手机号不能为空'}), 400
    
    # 只允许超级管理员18184005669重置密码
    SUPER_ADMIN = '18184005669'
    if admin_phone != SUPER_ADMIN:
        return jsonify({'error': '请联系办公室徐钟亿进行密码重置'}), 403
    
    # 禁止重置超级管理员自己的账号（防止被滥用）
    if phone == SUPER_ADMIN:
        return jsonify({'error': '超级管理员账号不能通过此方式重置密码，请联系系统管理员'}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # 验证管理员身份
        cur.execute("SELECT is_admin FROM users WHERE phone=%s", (admin_phone,))
        admin = cur.fetchone()
        if not admin or not admin['is_admin']:
            cur.close(); conn.close()
            return jsonify({'error': '无权限'}), 403
        
        # 重置为默认密码
        cur.execute("UPDATE users SET password='123456' WHERE phone=%s", (phone,))
        if cur.rowcount == 0:
            cur.close(); conn.close()
            return jsonify({'error': '用户不存在'}), 404
        
        conn.commit()
        cur.close(); conn.close()
        return jsonify({'success': True, 'message': '密码已重置为123456'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============= Render部署 =============
if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
